"""Utilities for importing requirements from tabular data sources."""

from __future__ import annotations

from dataclasses import dataclass, field
from enum import Enum
import csv
from pathlib import Path
import re
from typing import Any, Iterable, Sequence

from .model import (
    Priority,
    Requirement,
    RequirementType,
    Status,
    Verification,
    requirement_from_dict,
)

try:  # pragma: no cover - import guarded for environments without Excel support
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - fallback when openpyxl is missing
    load_workbook = None  # type: ignore[assignment]

__all__ = [
    "ImportFieldSpec",
    "RequirementImportConfiguration",
    "RequirementImportError",
    "RequirementImportIssue",
    "RequirementImportResult",
    "SequentialIDAllocator",
    "TabularDataset",
    "TabularFileFormat",
    "build_requirements",
    "detect_format",
    "importable_fields",
    "list_excel_sheets",
    "load_csv_dataset",
    "load_excel_dataset",
]


class RequirementImportError(Exception):
    """Raised when tabular data cannot be interpreted for import."""


class RequirementImportRowError(RequirementImportError):
    """Raised for problems specific to a single source row."""

    def __init__(self, message: str, *, field: str | None = None) -> None:
        super().__init__(message)
        self.field = field


class DuplicateIdentifierError(RequirementImportRowError):
    """Raised when a requirement id collides with existing data."""


class TabularFileFormat(str, Enum):
    """Supported spreadsheet formats."""

    CSV = "csv"
    EXCEL = "excel"


@dataclass(slots=True)
class TabularDataset:
    """Container for raw tabular values."""

    rows: list[list[Any]]
    _column_count: int = field(init=False, default=0)
    _header_row: list[Any] | None = field(init=False, default=None)

    def __post_init__(self) -> None:
        self._column_count = max((len(row) for row in self.rows), default=0)
        self._header_row = self.rows[0] if self.rows else None

    @property
    def column_count(self) -> int:
        return self._column_count

    @property
    def header(self) -> list[str] | None:
        if self._header_row is None:
            return None
        return [_stringify(cell) for cell in self._header_row]

    def column_names(self, *, use_header: bool) -> list[str]:
        names: list[str] = []
        header = self.header if use_header else None
        for index in range(self._column_count):
            label = ""
            if header and index < len(header):
                label = header[index].strip()
            if not label:
                label = f"Column {index + 1}"
            names.append(label)
        return names

    def iter_rows(self, *, skip_header: bool = False) -> Iterable[list[Any]]:
        start = 1 if skip_header and self._header_row is not None else 0
        for row in self.rows[start:]:
            yield row

    def row_count(self, *, skip_header: bool = False) -> int:
        if not self.rows:
            return 0
        return len(self.rows) - (1 if skip_header and self._header_row is not None else 0)


@dataclass(slots=True, frozen=True)
class ImportFieldSpec:
    """Describe a requirement field supported during import."""

    name: str
    required: bool = False
    enum: type[Enum] | None = None
    multi_value: bool = False
    synonyms: tuple[str, ...] = ()


importable_fields: tuple[ImportFieldSpec, ...] = (
    ImportFieldSpec("id", synonyms=("identifier", "req id", "requirement id")),
    ImportFieldSpec("title", synonyms=("name", "summary")),
    ImportFieldSpec(
        "statement",
        required=True,
        synonyms=("requirement", "text", "description"),
    ),
    ImportFieldSpec("type", enum=RequirementType, synonyms=("category",)),
    ImportFieldSpec("status", enum=Status),
    ImportFieldSpec("owner", synonyms=("assignee", "responsible")),
    ImportFieldSpec("priority", enum=Priority, synonyms=("importance",)),
    ImportFieldSpec("source", synonyms=("origin", "reference")),
    ImportFieldSpec("verification", enum=Verification),
    ImportFieldSpec("acceptance", synonyms=("acceptance criteria",)),
    ImportFieldSpec("conditions"),
    ImportFieldSpec("rationale", synonyms=("reason", "justification")),
    ImportFieldSpec("assumptions"),
    ImportFieldSpec("notes", synonyms=("comment", "comments")),
    ImportFieldSpec("labels", multi_value=True, synonyms=("tags", "categories")),
    ImportFieldSpec("approved_at"),
    ImportFieldSpec("modified_at"),
)

_FIELD_MAP = {spec.name: spec for spec in importable_fields}
_ENUM_FIELDS = {spec.name: spec.enum for spec in importable_fields if spec.enum}


@dataclass(slots=True)
class RequirementImportConfiguration:
    """User-selected mapping between columns and requirement fields."""

    mapping: dict[str, int | None]
    has_header: bool
    labels_separator: str = ","

    def __post_init__(self) -> None:
        normalized: dict[str, int | None] = {}
        for field, column in self.mapping.items():
            if field not in _FIELD_MAP:
                continue
            if column is None:
                normalized[field] = None
            else:
                normalized[field] = int(column)
        self.mapping = normalized
        spec = _FIELD_MAP.get("statement")
        if spec and spec.required and self.mapping.get("statement") is None:
            raise RequirementImportError("statement field must be mapped")


@dataclass(slots=True)
class RequirementImportIssue:
    """Description of a problem encountered while processing a row."""

    row: int
    field: str | None
    message: str


@dataclass(slots=True)
class RequirementImportResult:
    """Outcome of a bulk conversion of rows into requirements."""

    requirements: list[Requirement]
    issues: list[RequirementImportIssue]
    processed_rows: int
    imported_rows: int
    skipped_rows: int
    truncated: bool = False


class SequentialIDAllocator:
    """Allocate unique requirement identifiers while respecting existing ones."""

    def __init__(self, *, start: int, existing: Iterable[int] = ()) -> None:
        self._used: set[int] = {int(value) for value in existing}
        self._next = max(start, max(self._used, default=0) + 1)

    def clone(self) -> "SequentialIDAllocator":
        clone = SequentialIDAllocator(start=self._next)
        clone._used = set(self._used)
        clone._next = self._next
        return clone

    def reserve(self, value: int) -> None:
        if value in self._used:
            raise DuplicateIdentifierError(
                f"requirement id {value} already exists", field="id"
            )
        self._used.add(value)
        if value >= self._next:
            self._next = value + 1

    def allocate(self) -> int:
        candidate = self._next
        while candidate in self._used:
            candidate += 1
        self._used.add(candidate)
        self._next = candidate + 1
        return candidate


def detect_format(path: str | Path) -> TabularFileFormat:
    """Infer tabular format from file extension."""

    suffix = Path(path).suffix.lower()
    if suffix in {".csv", ".tsv"}:
        return TabularFileFormat.CSV
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return TabularFileFormat.EXCEL
    raise RequirementImportError(f"unsupported file format: {suffix or 'unknown'}")


def _normalize_delimiter(delimiter: str) -> str:
    if not delimiter:
        raise RequirementImportError("delimiter cannot be empty")
    if len(delimiter) != 1:
        raise RequirementImportError("delimiter must be a single character")
    return delimiter


def load_csv_dataset(path: str | Path, *, delimiter: str = ",") -> TabularDataset:
    """Load CSV/TSV file into a :class:`TabularDataset`."""

    norm_delim = _normalize_delimiter(delimiter)
    rows: list[list[Any]] = []
    with Path(path).open(encoding="utf-8") as fh:
        reader = csv.reader(fh, delimiter=norm_delim)
        for row in reader:
            rows.append([cell for cell in row])
    return TabularDataset(rows)


def list_excel_sheets(path: str | Path) -> list[str]:
    """Return sheet names available within an Excel workbook."""

    if load_workbook is None:
        raise RequirementImportError("openpyxl is not available")
    workbook = load_workbook(filename=Path(path), read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def load_excel_dataset(
    path: str | Path,
    *,
    sheet: str | None = None,
) -> TabularDataset:
    """Load Excel sheet into a :class:`TabularDataset`."""

    if load_workbook is None:
        raise RequirementImportError("openpyxl is not available")
    workbook = load_workbook(filename=Path(path), read_only=True, data_only=True)
    try:
        if sheet:
            if sheet not in workbook.sheetnames:
                raise RequirementImportError(f"sheet not found: {sheet}")
            worksheet = workbook[sheet]
        else:
            worksheet = workbook.active
        rows: list[list[Any]] = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append(list(row))
        return TabularDataset(rows)
    finally:
        workbook.close()


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def _is_blank_row(row: Sequence[Any]) -> bool:
    return all(_is_blank(cell) for cell in row)


def _parse_id(value: Any) -> int:
    if value is None:
        raise RequirementImportRowError("id cannot be empty", field="id")
    if isinstance(value, bool):
        raise RequirementImportRowError("id cannot be boolean", field="id")
    if isinstance(value, (int, float)):
        if isinstance(value, float):
            if not value.is_integer():
                raise RequirementImportRowError(
                    "id must be an integer value", field="id"
                )
            value = int(value)
        number = int(value)
        if number <= 0:
            raise RequirementImportRowError("id must be positive", field="id")
        return number
    text = str(value).strip()
    if not text:
        raise RequirementImportRowError("id cannot be empty", field="id")
    try:
        number = int(text)
    except ValueError as exc:
        raise RequirementImportRowError("id must be an integer", field="id") from exc
    if number <= 0:
        raise RequirementImportRowError("id must be positive", field="id")
    return number


def _normalize_enum(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, Enum):
        return value.value
    text = str(value).strip()
    if not text:
        return None
    normalized = re.sub(r"[\s-]+", "_", text.lower())
    return normalized


def _extract_field_value(
    row: Sequence[Any],
    *,
    index: int | None,
    field: str,
    separator: str,
) -> Any:
    if index is None:
        return None
    if index < 0 or index >= len(row):
        value: Any = None
    else:
        value = row[index]
    if field == "labels":
        text = _stringify(value)
        if not text:
            return []
        parts = re.split(r"[{}]+".format(re.escape(separator) + ";"), text)
        labels = [part.strip() for part in parts if part and part.strip()]
        return labels
    if field in _ENUM_FIELDS:
        return _normalize_enum(value)
    if field == "statement":
        return str(value) if value is not None else ""
    if isinstance(value, str):
        return value
    if value is None:
        return ""
    return str(value)


def _row_payload(
    row: Sequence[Any],
    config: RequirementImportConfiguration,
) -> tuple[dict[str, Any], int | None]:
    mapping = config.mapping
    payload: dict[str, Any] = {}
    candidate_id: int | None = None

    statement_index = mapping.get("statement")
    statement_value = _extract_field_value(
        row, index=statement_index, field="statement", separator=config.labels_separator
    )
    if not isinstance(statement_value, str) or not statement_value.strip():
        raise RequirementImportRowError("statement cannot be empty", field="statement")
    payload["statement"] = statement_value

    for field, index in mapping.items():
        if field in {"statement", "id"}:
            continue
        if index is None:
            continue
        value = _extract_field_value(
            row, index=index, field=field, separator=config.labels_separator
        )
        if field == "labels" and value == []:
            continue
        if value in (None, "") and field not in {"title", "notes"}:
            continue
        payload[field] = value

    if mapping.get("id") is not None:
        try:
            candidate_id = _parse_id(row[mapping["id"]])
        except IndexError:
            raise RequirementImportRowError("id column is missing", field="id")

    return payload, candidate_id


def build_requirements(
    dataset: TabularDataset,
    config: RequirementImportConfiguration,
    *,
    allocator: SequentialIDAllocator,
    max_rows: int | None = None,
) -> RequirementImportResult:
    """Convert a dataset into Requirement objects using ``config`` mapping."""

    requirements: list[Requirement] = []
    issues: list[RequirementImportIssue] = []
    processed = 0
    imported = 0
    skipped = 0
    truncated = False

    for row_index, row in enumerate(
        dataset.iter_rows(skip_header=config.has_header), start=1
    ):
        if max_rows is not None and processed >= max_rows:
            truncated = True
            break
        processed += 1
        if _is_blank_row(row):
            skipped += 1
            continue
        try:
            payload, candidate_id = _row_payload(row, config)
        except RequirementImportRowError as exc:
            issues.append(
                RequirementImportIssue(row=row_index, field=exc.field, message=str(exc))
            )
            continue
        try:
            if candidate_id is not None:
                allocator.reserve(candidate_id)
                payload["id"] = candidate_id
            else:
                payload["id"] = allocator.allocate()
        except RequirementImportRowError as exc:
            issues.append(
                RequirementImportIssue(row=row_index, field=exc.field, message=str(exc))
            )
            continue
        try:
            requirement = requirement_from_dict(payload)
        except (TypeError, ValueError) as exc:
            issues.append(
                RequirementImportIssue(row=row_index, field=None, message=str(exc))
            )
            continue
        imported += 1
        requirements.append(requirement)

    return RequirementImportResult(
        requirements=requirements,
        issues=issues,
        processed_rows=processed,
        imported_rows=imported,
        skipped_rows=skipped,
        truncated=truncated,
    )

